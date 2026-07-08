"""Split a huge .xlsx into smaller files without loading it into memory.

Usage: python split_xlsx.py big.xlsx --rows-per-file 100000
"""
import argparse
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook, Workbook


class XlsxSplitter:
    def __init__(self, src: Path, rows_per_file: int, keep_header: bool = True):
        self.src = src
        self.rows_per_file = rows_per_file
        self.keep_header = keep_header

    def _rows(self) -> Iterator[tuple]:
        # read_only=True streams rows from the zip; constant memory
        wb = load_workbook(self.src, read_only=True, data_only=True)
        ws = wb.active
        yield from ws.iter_rows(values_only=True)
        wb.close()

    def split(self) -> list[Path]:
        rows = self._rows()
        header = next(rows) if self.keep_header else None
        outputs = []

        wb = ws = None
        count = 0
        part = 0

        def new_part():
            nonlocal wb, ws, part, count
            self._flush(wb, part, outputs)
            part += 1
            wb = Workbook(write_only=True)  # streaming writer
            ws = wb.create_sheet()
            if header:
                ws.append(header)
            count = 0

        new_part()
        for row in rows:
            ws.append(row)
            count += 1
            if count >= self.rows_per_file:
                new_part()
        self._flush(wb, part, outputs)
        return outputs

    def _flush(self, wb, part: int, outputs: list[Path]):
        if wb is None:
            return
        out = self.src.with_name(f"{self.src.stem}_part{part:03d}.xlsx")
        wb.save(out)
        outputs.append(out)


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("file", type=Path)
    p.add_argument("--rows-per-file", type=int, default=100_000)
    p.add_argument("--no-header", action="store_true")
    args = p.parse_args()

    for f in XlsxSplitter(args.file, args.rows_per_file, not args.no_header).split():
        print(f)
